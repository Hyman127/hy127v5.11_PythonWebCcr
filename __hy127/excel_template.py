
""" 2025.11.25 作者: 郑广学 hy127.cn
vsersion: 0.0.2
excel_template excel模板引擎 实现类似word邮件合并的模板套打组件

针对 WPS/Office 自定义默认行高列宽的完美修复版 + 图片居中插入（OneCellAnchor 方案）
新增：区域复制粘贴、区域模板填充功能

uv add openpyxl

该模块提供了一个Excel模板引擎，可以将数据填充到Excel模板中，支持简单字段替换、
集合数据展开、图片插入等功能。类似于Word的邮件合并功能，但专门用于Excel文件。

基本用法：
1. 创建模板：在Excel中使用 {{字段名}} 的格式标记占位符
2. 准备数据：准备字典或对象形式的数据
3. 渲染模板：使用ExcelTemplate类将数据填充到模板中

Excel模板引擎类
    
    主要功能：
    1. 字段替换：将 {{字段名}} 替换为实际值
    2. 集合展开：将列表数据展开成多行
    3. 图片插入：在指定位置插入图片并自动调整大小和居中
    4. 格式保留：保持原有单元格格式（字体、边框、颜色等）
    5. 尺寸适配：自动适配不同Excel软件的默认行列尺寸
    6. 区域复制：复制指定区域到目标位置（带格式、行高列宽、合并单元格）
    7. 区域模板填充：对指定区域进行数据模板填充
    
    支持的标签格式：
    - {{字段名}} : 普通字段替换
    - {{@img:字段名}} : 图片插入标签
    - {{字段名.子字段}} : 嵌套字段访问
    - {{@ext:字段名.子字段}} : 集合数据扩展行模式
    
示例：
    # 简单数据渲染
    template = ExcelTemplate("template.xlsx")
    data = {
        "name": "张三",
        "age": 25,
        "department": "技术部"
    }
    template.save_as_by_template("output.xlsx", data)
    
    # 集合数据渲染（表格形式）
    data = {
        "title": "员工列表",
        "employees": [
            {"name": "张三", "age": 25},
            {"name": "李四", "age": 30}
        ]
    }
    template.save_as_by_template("employees.xlsx", data)
    
    # 插入图片
    data = {
        "photo": "path/to/image.jpg"  # 或者PIL图像对象或者bytes数据
    }
    # 在模板中使用 {{@img:photo}} 标记图片位置
    template.save_as_by_template("with_photo.xlsx", data)
    
    # 区域复制粘贴
    ExcelTemplate.copy_range(
        workbook=wb,
        source_sheet="Sheet1",
        source_range="A1:D10",
        target_sheet="Sheet2",
        target_start="B5"
    )
    
    # 区域模板填充
    data = {"name": "张三", "score": 95}
    ExcelTemplate.fill_range_with_template(
        workbook=wb,
        sheet_name="Sheet1",
        range_address="A1:C5",
        data=data
    )

"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, EMU_to_pixels
from openpyxl.worksheet.cell_range import CellRange
import re
import os
import logging
from typing import Any, Dict, List, Union, Tuple
from copy import copy, deepcopy
from io import BytesIO

logger = logging.getLogger(__name__)


class ExcelTemplate:
    """
    Excel模板引擎类
    
    主要功能：
    1. 字段替换：将 {{字段名}} 替换为实际值
    2. 集合展开：将列表数据展开成多行
    3. 图片插入：在指定位置插入图片并自动调整大小和居中
    4. 格式保留：保持原有单元格格式（字体、边框、颜色等）
    5. 尺寸适配：自动适配不同Excel软件的默认行列尺寸
    6. 区域复制：复制指定区域到目标位置（带格式、行高列宽、合并单元格）
    7. 区域模板填充：对指定区域进行数据模板填充
    
    支持的标签格式：
    - {{字段名}} : 普通字段替换
    - {{@img:字段名}} : 图片插入标签
    - {{字段名.子字段}} : 嵌套字段访问
    - {{@ext:字段名.子字段}} : 集合数据扩展行模式
    """
    
    # ========================================================================
    # 像素换算系数
    # ========================================================================
    COL_WIDTH_SCALE = 8.0  
    ROW_HEIGHT_SCALE = 1.333 
    FALLBACK_COL_WIDTH = 8.43
    FALLBACK_ROW_HEIGHT = 15.0

    def __init__(self, template_path: str):
        """
        初始化模板引擎
        
        Args:
            template_path (str): Excel模板文件路径
        """
        self.template_path = template_path
        self.workbook = None
        self.tag_pattern = re.compile(r'\{\{(@img:|@ext:)?(\w+(?:\.\w+)*)\}\}')
        
    def 填充模板并保存文件(self, output_path: str, value: Union[Dict[str, Any], object]):
        """
        根据模板生成新的Excel文件并保存
        
        Args:
            output_path (str): 输出文件路径
            value (Union[Dict[str, Any], object]): 数据源，可以是字典或对象
        """
        self.workbook = load_workbook(self.template_path)
        data = self._to_dict(value)
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            self._process_sheet(sheet, data)
        self.workbook.save(output_path)
    
    def _to_dict(self, obj):
        """将对象转换为字典格式"""
        if isinstance(obj, dict): return obj
        return obj.__dict__ if hasattr(obj, '__dict__') else {}

    def _process_sheet(self, sheet: Worksheet, data: Dict):
        """
        处理工作表数据填充
        
        Args:
            sheet (Worksheet): 工作表对象
            data (Dict): 数据字典
        """
        collection_fields = {k: v for k, v in data.items() if isinstance(v, (list, tuple))}
        if not collection_fields:
            self._replace_all_tags(sheet, data)
            return
        
        rows_with_collections = self._find_collection_rows(sheet, collection_fields)
        for row_info in sorted(rows_with_collections, key=lambda x: x['row'], reverse=True):
            if row_info['extend_mode']:
                self._expand_collection_row(sheet, row_info, data)
            else:
                self._fill_collection_in_place(sheet, row_info, data)
        self._replace_all_tags(sheet, data)

    # =========================================================================
    # 核心修复区域：尺寸计算
    # =========================================================================
    
    def _get_sheet_defaults(self, sheet: Worksheet) -> Tuple[float, float]:
        """获取当前 Sheet 的全局默认宽、高"""
        default_w = self.FALLBACK_COL_WIDTH
        if sheet.sheet_format and sheet.sheet_format.defaultColWidth:
            default_w = sheet.sheet_format.defaultColWidth
        
        default_h = self.FALLBACK_ROW_HEIGHT
        if sheet.sheet_format and sheet.sheet_format.defaultRowHeight:
            default_h = sheet.sheet_format.defaultRowHeight
            
        return default_w, default_h

    def _calculate_merged_cell_size(self, sheet: Worksheet, r) -> Tuple[float, float]:
        """计算合并单元格区域的像素尺寸"""
        def_w, def_h = self._get_sheet_defaults(sheet)
        
        total_w = 0
        for col_idx in range(r.min_col, r.max_col + 1):
            col_letter = get_column_letter(col_idx)
            dim = sheet.column_dimensions.get(col_letter)
            w = dim.width if dim and dim.width is not None else def_w
            total_w += w * self.COL_WIDTH_SCALE
            
        total_h = 0
        for row_idx in range(r.min_row, r.max_row + 1):
            dim = sheet.row_dimensions.get(row_idx)
            h = dim.height if dim and dim.height is not None else def_h
            total_h += h * self.ROW_HEIGHT_SCALE
            
        return total_w, total_h

    def _calculate_single_cell_size(self, sheet: Worksheet, cell) -> Tuple[float, float]:
        def_w, def_h = self._get_sheet_defaults(sheet)
        
        col_letter = get_column_letter(cell.column)
        dim_c = sheet.column_dimensions.get(col_letter)
        w = dim_c.width if dim_c and dim_c.width is not None else def_w
        
        dim_r = sheet.row_dimensions.get(cell.row)
        h = dim_r.height if dim_r and dim_r.height is not None else def_h
        
        return (w * self.COL_WIDTH_SCALE, h * self.ROW_HEIGHT_SCALE)

    # =========================================================================
    # 核心改进：图片居中插入（使用 OneCellAnchor）
    # =========================================================================
    def _insert_image_to_cell(self, sheet, cell, image_data):
        """插入图片并在单元格中居中（使用 OneCellAnchor 方案）"""
        try:
            # 创建图片对象
            img = self._create_image_object(image_data)
            
            # 检查是否为合并单元格
            merged_range = self._get_merged_range(sheet, cell)
            
            if merged_range:
                # 合并单元格：计算整个区域的尺寸
                cell_w, cell_h = self._calculate_merged_cell_size(sheet, merged_range)
                anchor_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            else:
                # 单个单元格：计算单元格尺寸
                cell_w, cell_h = self._calculate_single_cell_size(sheet, cell)
                anchor_cell = cell
            
            # 调整图片大小以适应单元格（保持宽高比，留 2px padding）
            if cell_w > 0 and cell_h > 0:
                self._resize_image_to_fit(img, cell_w, cell_h, padding=2)
            
            # ==================== 居中计算 ====================
            # 图片缩放后的实际尺寸（像素）
            img_w = img.width
            img_h = img.height
            
            # 计算居中偏移量（像素）
            offset_x = max(0, (cell_w - img_w) / 2)
            offset_y = max(0, (cell_h - img_h) / 2)
            
            # 转换为 EMU 单位
            col_off_emu = pixels_to_EMU(offset_x)
            row_off_emu = pixels_to_EMU(offset_y)
            
            # ==================== 使用 OneCellAnchor（关键修复）====================
            # 创建起始锚点
            from_marker = AnchorMarker(
                col=anchor_cell.column - 1,
                row=anchor_cell.row - 1,
                colOff=col_off_emu,
                rowOff=row_off_emu
            )
            
            # 创建图片尺寸对象（EMU 单位）
            ext = XDRPositiveSize2D(
                cx=pixels_to_EMU(img_w),  # 图片宽度
                cy=pixels_to_EMU(img_h)   # 图片高度
            )
            
            # 使用 OneCellAnchor（只指定起始位置和尺寸）
            anchor = OneCellAnchor(_from=from_marker, ext=ext)
            
            # 添加图片到工作表
            sheet.add_image(img, anchor)
            
            # 清除单元格内容
            cell.value = None
            
        except Exception as e:
            logger.exception("插入图片失败")
            cell.value = f"Img Error"

    def _resize_image_to_fit(self, img, max_w, max_h, padding=2):
        """调整图片大小以适应单元格（保持宽高比）"""
        target_w = max(max_w - padding * 2, 1)
        target_h = max(max_h - padding * 2, 1)
        
        ratio_img = img.width / img.height
        ratio_cell = target_w / target_h
        
        if ratio_img > ratio_cell:
            # 图片较宽，受限于宽度
            img.width = target_w
            img.height = target_w / ratio_img
        else:
            # 图片较高，受限于高度
            img.height = target_h
            img.width = target_h * ratio_img

    # =========================================================================
    # 辅助方法
    # =========================================================================
    
    def _get_merged_range(self, sheet, cell):
        """检查单元格是否为合并单元格"""
        for r in sheet.merged_cells.ranges:
            if cell.coordinate in r: 
                return r
        return None
        
    def _replace_all_tags(self, sheet, data):
        """替换所有非集合字段的标签"""
        simple_data = {k: v for k, v in data.items() if not isinstance(v, (list, tuple))}
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if self._should_insert_image(cell.value):
                        self._process_image_cell(sheet, cell, cell.value, simple_data)
                    else:
                        # 直接替换，保留原始数据类型
                        new_value = self._replace_tags_in_string(cell.value, simple_data)
                        cell.value = new_value

    def _should_insert_image(self, text):
        """判断是否为图片标签"""
        return bool(re.search(r'\{\{@img:', text))

    def _process_image_cell(self, sheet, cell, text, data, col_name=None):
        """处理图片单元格"""
        matches = self.tag_pattern.findall(text)
        for prefix, tag in matches:
            if prefix == '@img:':
                val = self._get_nested_value(data, tag)
                if val:
                    self._insert_image_to_cell(sheet, cell, val)
                    return
        cell.value = None

    def _replace_tags_in_string(self, text, data, col_name=None):
        """
        替换字符串中的标签
        
        关键修复：
        1. 如果整个单元格只有一个标签，返回原始类型（保留数字、日期等类型）
        2. 如果是混合文本，才转换为字符串
        3. 支持 @ext: 前缀标签的替换
        """
        # 检查是否整个单元格就是一个标签
        match = self.tag_pattern.fullmatch(text)
        if match:
            prefix, tag = match.groups()
            # 跳过图片标签
            if prefix == '@img:':
                return text
            # 返回原始值（保留数据类型）
            val = self._get_nested_value(data, tag)
            return val if val is not None else ''
        
        # 混合文本情况：需要拼接字符串
        def repl(m):
            prefix, tag = m.groups()
            # 跳过图片标签
            if prefix == '@img:':
                return m.group(0)
            # @ext: 和普通标签都正常替换
            val = self._get_nested_value(data, tag)
            return str(val) if val is not None else ''
        
        return self.tag_pattern.sub(repl, text)

    def _get_nested_value(self, data, tag):
        """获取嵌套属性值（支持 a.b.c 格式）"""
        try:
            v = data
            for k in tag.split('.'):
                if isinstance(v, dict):
                    v = v.get(k)
                elif isinstance(v, (list, tuple)):
                    # 如果是列表，尝试将k转换为索引
                    try:
                        idx = int(k)
                        v = v[idx] if 0 <= idx < len(v) else None
                    except (ValueError, IndexError):
                        v = None
                else:
                    v = getattr(v, k, None)
                
                if v is None:
                    break
            return v
        except (AttributeError, TypeError, ValueError, IndexError): 
            return None

    def _create_image_object(self, data):
        """从路径/PIL/bytes创建图片对象"""
        if isinstance(data, str): 
            return XLImage(data)
        try:
            from PIL import Image
            if isinstance(data, Image.Image):
                b = BytesIO()
                data.save(b, format='PNG')
                b.seek(0)
                return XLImage(b)
        except ImportError: 
            pass
        if isinstance(data, (bytes, BytesIO)):
            if isinstance(data, bytes): 
                data = BytesIO(data)
            data.seek(0)
            return XLImage(data)
        raise ValueError("Unknown image type")

    # =========================================================================
    # 集合数据处理
    # =========================================================================
    
    def _find_collection_rows(self, sheet, collection_fields):
        """查找包含集合字段的行"""
        result = []
        processed_rows = set()
        for row_idx in range(1, sheet.max_row + 1):
            if row_idx in processed_rows: 
                continue
            row_tags = []
            has_extend = False
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row_idx, col_idx).value
                if isinstance(val, str):
                    matches = self.tag_pattern.findall(val)
                    for prefix, tag in matches:
                        # 存储实际标签名（去除前缀）
                        row_tags.append({'tag': tag, 'has_ext_prefix': prefix == '@ext:'})
                        if prefix == '@ext:': 
                            has_extend = True
            
            for name, col_data in collection_fields.items():
                if not col_data: 
                    continue
                # 检查是否有标签以 集合名. 开头
                if any(t['tag'].startswith(name + '.') for t in row_tags):
                    result.append({
                        'row': row_idx, 
                        'collection_name': name,
                        'collection_data': col_data, 
                        'extend_mode': has_extend
                    })
                    processed_rows.add(row_idx)
                    break
        return result

    def _fill_collection_in_place(self, sheet, row_info, data):
        """不扩展行，直接覆盖现有行"""
        start = row_info['row']
        items = row_info['collection_data']
        name = row_info['collection_name']
        
        if not items:
            for col in range(1, sheet.max_column + 1): 
                sheet.cell(start, col).value = None
            return
            
        original = self._save_row_template(sheet, start)
        for idx, item in enumerate(items):
            current = start + idx
            if current > sheet.max_row: 
                break
            full_data = data.copy()
            full_data[name] = self._to_dict(item)
            self._restore_row_template(sheet, current, original, full_data, name)

    def _expand_collection_row(self, sheet, row_info, data):
        """扩展模式：插入新行来填充集合"""
        start = row_info['row']
        items = row_info['collection_data']
        name = row_info['collection_name']
        
        if not items:
            sheet.delete_rows(start, 1)
            return
            
        original = self._save_row_template(sheet, start)
        
        if len(items) > 1:
            sheet.insert_rows(start + 1, len(items) - 1)
            
        for idx, item in enumerate(items):
            full_data = data.copy()
            full_data[name] = self._to_dict(item)
            self._restore_row_template(sheet, start + idx, original, full_data, name)

    def _save_row_template(self, sheet, row_idx):
        """保存行的格式和内容作为模板（包括行高）"""
        template = []
        
        # 保存行高
        row_dim = sheet.row_dimensions.get(row_idx)
        row_height = row_dim.height if row_dim and row_dim.height is not None else None
        
        # 保存每个单元格的信息
        for col in range(1, sheet.max_column + 1):
            c = sheet.cell(row_idx, col)
            template.append({
                'value': c.value,
                'font': copy(c.font), 
                'border': copy(c.border),
                'fill': copy(c.fill), 
                'number_format': c.number_format,
                'alignment': copy(c.alignment)
            })
        
        # 返回模板（包含行高信息）
        return {
            'cells': template,
            'row_height': row_height
        }

    def _restore_row_template(self, sheet, row_idx, template, data, col_name=None):
        """根据模板恢复行（保留数据类型和行高）"""
        cells_template = template['cells']
        row_height = template['row_height']
        
        # 恢复行高
        if row_height is not None:
            sheet.row_dimensions[row_idx].height = row_height
        
        # 恢复每个单元格
        for i, info in enumerate(cells_template, 1):
            cell = sheet.cell(row_idx, i)
            val = info['value']
            
            if val:
                if isinstance(val, str) and self._should_insert_image(val):
                    # 处理图片
                    self._process_image_cell(sheet, cell, val, data, col_name)
                elif isinstance(val, str) and self.tag_pattern.search(val):
                    # 替换标签（保留原始数据类型）
                    replaced_val = self._replace_tags_in_string(val, data, col_name)
                    cell.value = replaced_val
                else:
                    # 非标签内容，直接赋值
                    cell.value = val
            
            # 恢复单元格格式
            if info['font']: 
                cell.font = copy(info['font'])
            if info['border']: 
                cell.border = copy(info['border'])
            if info['fill']: 
                cell.fill = copy(info['fill'])
            cell.number_format = info['number_format']
            if info['alignment']: 
                cell.alignment = copy(info['alignment'])

    # =========================================================================
    # 新增静态方法：区域复制粘贴（带格式、行高列宽、合并单元格）
    # =========================================================================
    
    @staticmethod
    def 复制区域(
        workbook,
        source_sheet: Union[str, Worksheet],
        source_range: str,
        target_sheet: Union[str, Worksheet],
        target_start: str
    ):
        """
        复制指定区域到目标位置，保留所有格式、行高、列宽和合并单元格
        
        Args:
            workbook: openpyxl Workbook 对象
            source_sheet: 源工作表名称或对象
            source_range: 源区域地址，如 "A1:D10"
            target_sheet: 目标工作表名称或对象
            target_start: 目标起始单元格，如 "B5"
        """
        # 获取工作表对象
        if isinstance(source_sheet, str):
            source_sheet = workbook[source_sheet]
        if isinstance(target_sheet, str):
            target_sheet = workbook[target_sheet]
        
        # 解析源区域范围
        min_col, min_row, max_col, max_row = range_boundaries(source_range)
        
        # 解析目标起始位置
        from openpyxl.utils.cell import coordinate_to_tuple
        target_row, target_col = coordinate_to_tuple(target_start)
        
        # 计算行列偏移量
        row_offset = target_row - min_row
        col_offset = target_col - min_col
        
        # 1. 复制单元格内容和格式
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                source_cell = source_sheet.cell(row_idx, col_idx)
                target_cell = target_sheet.cell(row_idx + row_offset, col_idx + col_offset)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 复制格式
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        
        # 2. 复制行高
        for row_idx in range(min_row, max_row + 1):
            source_row_dim = source_sheet.row_dimensions.get(row_idx)
            if source_row_dim and source_row_dim.height is not None:
                target_row_idx = row_idx + row_offset
                target_sheet.row_dimensions[target_row_idx].height = source_row_dim.height
        
        # 3. 复制列宽
        for col_idx in range(min_col, max_col + 1):
            source_col_letter = get_column_letter(col_idx)
            source_col_dim = source_sheet.column_dimensions.get(source_col_letter)
            
            if source_col_dim and source_col_dim.width is not None:
                target_col_letter = get_column_letter(col_idx + col_offset)
                target_sheet.column_dimensions[target_col_letter].width = source_col_dim.width
        
        # 4. 复制合并单元格（关键修复：先收集再处理）
        merged_ranges_to_copy = []
        for merged_range in source_sheet.merged_cells.ranges:
            # 检查合并单元格是否在源区域内
            if (merged_range.min_row >= min_row and merged_range.max_row <= max_row and
                merged_range.min_col >= min_col and merged_range.max_col <= max_col):
                merged_ranges_to_copy.append(merged_range)
        
        # 统一创建目标合并单元格
        for merged_range in merged_ranges_to_copy:
            new_min_row = merged_range.min_row + row_offset
            new_max_row = merged_range.max_row + row_offset
            new_min_col = merged_range.min_col + col_offset
            new_max_col = merged_range.max_col + col_offset
            
            # 创建新的合并单元格
            target_sheet.merge_cells(
                start_row=new_min_row,
                start_column=new_min_col,
                end_row=new_max_row,
                end_column=new_max_col
            )

    # =========================================================================
    # 新增静态方法：对指定区域实施模板填充
    # =========================================================================
    
    @staticmethod
    def 区域模板填充(
        workbook,
        sheet_name: Union[str, Worksheet],
        range_address: str,
        data: Union[Dict[str, Any], object]
    ):
        """
        对指定表的指定区域实施模板填充
        
        在指定区域内查找 {{字段名}} 标签并替换为数据值，保留所有格式
        
        Args:
            workbook: openpyxl Workbook 对象
            sheet_name: 工作表名称或对象
            range_address: 区域地址，如 "A1:D10"
            data: 数据字典或对象
        
        示例:
            wb = load_workbook("template.xlsx")
            data = {"name": "张三", "score": 95, "grade": "优秀"}
            ExcelTemplate.fill_range_with_template(
                workbook=wb,
                sheet_name="Sheet1",
                range_address="A1:C5",
                data=data
            )
            wb.save("output.xlsx")
        """
        # 获取工作表对象
        if isinstance(sheet_name, str):
            sheet = workbook[sheet_name]
        else:
            sheet = sheet_name
        
        # 转换数据为字典格式
        if not isinstance(data, dict):
            data = data.__dict__ if hasattr(data, '__dict__') else {}
        
        # 解析区域范围
        min_col, min_row, max_col, max_row = range_boundaries(range_address)
        
        # 创建标签匹配模式
        tag_pattern = re.compile(r'\{\{(@img:|@ext:)?(\w+(?:\.\w+)*)\}\}')
        
        # 遍历区域内的所有单元格
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                
                # 只处理字符串类型的单元格值
                if isinstance(cell.value, str):
                    # 检查是否包含标签
                    if '{{' in cell.value and '}}' in cell.value:
                        # 检查是否为图片标签
                        if '{{@img:' in cell.value:
                            # 处理图片标签
                            matches = tag_pattern.findall(cell.value)
                            for prefix, tag in matches:
                                if prefix == '@img:':
                                    val = ExcelTemplate._static_get_nested_value(data, tag)
                                    if val:
                                        try:
                                            # 创建临时实例来使用图片插入功能
                                            temp_instance = ExcelTemplate.__new__(ExcelTemplate)
                                            temp_instance._insert_image_to_cell(sheet, cell, val)
                                        except Exception as e:
                                            logger.exception("插入图片失败")
                                            cell.value = "Img Error"
                                    else:
                                        cell.value = None
                        else:
                            # 处理普通文本标签
                            cell.value = ExcelTemplate._static_replace_tags(
                                cell.value, data, tag_pattern
                            )
    
    @staticmethod
    def _static_get_nested_value(data: Dict, tag: str):
        """静态方法：获取嵌套属性值"""
        try:
            v = data
            for k in tag.split('.'):
                if isinstance(v, dict):
                    v = v.get(k)
                elif isinstance(v, (list, tuple)):
                    try:
                        idx = int(k)
                        v = v[idx] if 0 <= idx < len(v) else None
                    except (ValueError, IndexError):
                        v = None
                else:
                    v = getattr(v, k, None)
                
                if v is None:
                    break
            return v
        except (AttributeError, TypeError, ValueError, IndexError):
            return None
    
    @staticmethod
    def _static_replace_tags(text: str, data: Dict, tag_pattern):
        """静态方法：替换文本中的标签"""
        # 检查是否整个单元格就是一个标签
        match = tag_pattern.fullmatch(text)
        if match:
            prefix, tag = match.groups()
            if prefix == '@img:':
                return text
            val = ExcelTemplate._static_get_nested_value(data, tag)
            return val if val is not None else ''
        
        # 混合文本情况
        def repl(m):
            prefix, tag = m.groups()
            if prefix == '@img:':
                return m.group(0)
            val = ExcelTemplate._static_get_nested_value(data, tag)
            return str(val) if val is not None else ''
        
        return tag_pattern.sub(repl, text)


# =========================================================================
# 使用示例
# =========================================================================

if __name__ == "__main__":
    from openpyxl import Workbook
    
    # 示例1：基本模板填充
    print("=" * 60)
    print("示例1：基本模板填充")
    print("=" * 60)
    
    template = ExcelTemplate("template.xlsx")
    data = {
        "name": "张三",
        "age": 25,
        "department": "技术部",
        "salary": 8000.50
    }
    template.填充模板并保存文件("output.xlsx", data)
    print("✓ 基本模板填充完成: output.xlsx")
    
    # 示例2：集合数据渲染
    print("\n" + "=" * 60)
    print("示例2：集合数据渲染")
    print("=" * 60)
    
    data_with_list = {
        "title": "员工列表",
        "date": "2025-11-25",
        "employees": [
            {"name": "张三", "age": 25, "dept": "技术部"},
            {"name": "李四", "age": 30, "dept": "市场部"},
            {"name": "王五", "age": 28, "dept": "人事部"}
        ]
    }
    template.填充模板并保存文件("employees.xlsx", data_with_list)
    print("✓ 集合数据渲染完成: employees.xlsx")
    
    # 示例3：区域复制粘贴
    print("\n" + "=" * 60)
    print("示例3：区域复制粘贴")
    print("=" * 60)
    
    wb = load_workbook("source.xlsx")
    ExcelTemplate.复制区域(
        workbook=wb,
        source_sheet="Sheet1",
        source_range="A1:D10",
        target_sheet="Sheet2",
        target_start="B5"
    )
    wb.save("copied.xlsx")
    print("✓ 区域复制完成: copied.xlsx")
    print("  源区域: Sheet1!A1:D10")
    print("  目标位置: Sheet2!B5")
    
    # 示例4：区域模板填充
    print("\n" + "=" * 60)
    print("示例4：区域模板填充")
    print("=" * 60)
    
    wb = load_workbook("template.xlsx")
    data = {
        "student_name": "李明",
        "student_id": "20250001",
        "score": 95,
        "grade": "优秀",
        "teacher": "王老师"
    }
    ExcelTemplate.区域模板填充(
        workbook=wb,
        sheet_name="Sheet1",
        range_address="A1:E10",
        data=data
    )
    wb.save("filled_range.xlsx")
    print("✓ 区域模板填充完成: filled_range.xlsx")
    print("  填充区域: A1:E10")
    print(f"  填充数据: {data}")
