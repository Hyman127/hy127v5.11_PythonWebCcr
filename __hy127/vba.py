
# vba.py
"""
封装兼容vba的一些函数
创建者：郑广学 2025-10-6 vbayyds.com
完全启用需要安装包: uv add pandas pywin32 psutil

"""
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, timezone,timedelta
import re
import os
def round(数值, 小数位=0) -> float:
    """
    按Excel函数规则四舍五入。
    
    参数:
        number (float or str): 需要四舍五入的数字，可以是浮点数或字符串形式的数字。
        ndigits (int): 保留的小数位数，默认为 0，表示四舍五入到整数。
        
    返回:
        float: 四舍五入后的浮点数结果。
    """
    # 转为 Decimal 避免浮点误差
    d = Decimal(str(数值))
    # 使用 ROUND_HALF_UP（传统四舍五入）
    rounded = d.quantize(Decimal('0.1') ** 小数位, rounding=ROUND_HALF_UP)
    return float(rounded)

def val(s) -> float:
    """
    类似 VBA 的 Val 函数，增强支持布尔值：
    - True  → 1
    - False → 0
    - 字符串：从开头提取数字（支持 +/-、小数）
    - 无法识别的值（None、空串、无效文本等）→ 0
    """
    # 1. 先处理布尔值（必须在类型检查前，因为 bool 是 int 子类）
    if s is True:
        return 1
    if s is False:
        return 0
    
    # 2. 处理 None 和空值
    if s is None:
        return 0
    
    # 3. 转为字符串（但注意：bool 已处理，避免 True → "True"）
    if not isinstance(s, str):
        s = str(s)
    
    s = s.strip()
    if not s:
        return 0
    
    # 4. 用正则提取开头的有效数字
    match = re.match(r'^[+-]?(\d+\.?\d*|\.\d+)', s)
    if match:
        num_str = match.group()
        # 过滤无效模式如 "+", "-", ".", "+.", "-."
        if num_str in ('+', '-', '.', '+.', '-.') or not num_str:
            return 0
        try:
            return float(num_str) 
        except ValueError:
            return 0
    else:
        return 0



def str_to_datetime(value, format=None) -> datetime | None:
    """
    将字符串转换为datetime对象
    
    Args:
        value: 输入字符串
        format: 可选，字符串日期格式（如 "%y%m%d", "%Y-%m-%d"）
                若未提供，则尝试多种常见格式
    
    Returns:
        datetime: 转换后的datetime对象，无法转换返回None
    """
    if not isinstance(value, str):
        return None
        
    value = value.strip()
    if not value:
        return None

    dt = None
    if format is not None:
        # 使用用户指定的格式
        try:
            dt = datetime.strptime(value, format)
        except (ValueError, TypeError):
            return None
    else:
        # 自动尝试常见格式
        common_formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y%m%d",
            "%y%m%d",      # 支持 yymmdd
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in common_formats:
            try:
                dt = datetime.strptime(value, fmt)
                break
            except ValueError:
                continue
                
    return dt

def str_to_date(value, format=None) -> datetime | None:
    return str_to_datetime(value, format).date()
def cdate(value, format=None) -> float:
    """
    将各种输入转换为 Excel 兼容的日期数值（float）
    
    Args:
        value: 输入值（str, datetime, date, int, float, None）
        format: 可选，字符串日期格式（如 "%y%m%d", "%Y-%m-%d"）
                若未提供，则尝试多种常见格式
    
    Returns:
        float: Excel 日期数值（1900 日期系统），无效输入返回 0.0
    """
    # 1. 如果是数字，直接返回（假设已是 Excel 日期数值）
    if isinstance(value, (int, float)):
        return float(value) if value >= 0 else 0.0

    # 2. 处理 None 或空字符串
    if value is None:
        return 0.0
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return 0.0

    # 3. 尝试转换为 datetime 对象
    dt = None

    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    elif isinstance(value, str):
        dt = str_to_date(value, format)

    # 4. 仍无法解析，返回 0
    if dt is None:
        return 0.0

    # 5. 处理时区：转为本地时间
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)

    # 6. 转为 Excel 日期数值（1900 系统）
    excel_base = datetime(1899, 12, 30)
    delta = dt - excel_base
    excel_days = (
        delta.days 
        + delta.seconds / 86400.0 
        + delta.microseconds / 86400000000.0
    )
    return excel_days

def to_datetime(value, format=None) -> datetime | None:
    """
    将各种输入转换为Python datetime对象
    
    Args:
        value: 输入值（str, datetime, date, int, float, None）
        format: 可选，字符串日期格式（如 "%y%m%d", "%Y-%m-%d"）
                若未提供，则尝试多种常见格式
    
    Returns:
        datetime: Python datetime对象，无效输入返回 None
    """
    # 1. 如果已经是 datetime 对象，直接返回
    if isinstance(value, datetime):
        # 处理时区：转为本地时间
        if value.tzinfo is not None:
            value = value.astimezone().replace(tzinfo=None)
        return value

    # 2. 如果是 date 对象，转换为 datetime
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    # 3. 如果是数字，视为 Excel 日期数值
    if isinstance(value, (int, float)):
        if value < 0:
            return None
        # Excel 日期数值转 datetime（1900 系统）
        excel_base = datetime(1899, 12, 30)
        dt = excel_base + timedelta(days=value)
        return dt

    # 4. 处理 None 或空字符串
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

    # 5. 如果是字符串，使用 str_to_date 转换
    if isinstance(value, str):
        dt = str_to_date(value, format)
        return dt

    # 6. 其他情况无法处理，返回 None
    return None

def to_date(value, format=None) -> date | None:
    #将各种输入转换为Python date对象
    dt=to_datetime(value, format)
    if isinstance(dt,datetime):
        return dt.date()
    else:
        return dt

def IIF(条件, 真值, 假值):
    return 真值 if 条件 else 假值

def isEmpty(变量):
    return 变量 is None or 变量 == ""

def isNumberic(字符串):
    """使用正则表达式判断是否数值"""
    # 匹配整数、小数、负数、科学计数法
    模式 = r'^[-+]?[0-9]*\.?[0-9]+([eE][-+]?[0-9]+)?$'
    return bool(re.match(模式, 字符串))

def like(文本, 模式):
    """
    模拟VBA的Like运算符
    
    通配符：
        * - 匹配任意多个字符（包括0个）
        ? - 匹配单个字符
        # - 匹配单个数字(0-9)
        [abc] - 匹配括号中的任意一个字符
        [!abc] - 不匹配括号中的字符
    
    示例：
        like("Hello", "H*")      → True
        like("Hello", "H?llo")   → True
        like("123", "###")       → True
    """
    if 文本 is None or 模式 is None:
        return False
    
    # 转换VBA Like模式为正则表达式
    正则模式 = ""
    i = 0
    
    while i < len(模式):
        字符 = 模式[i]
        
        if 字符 == '*':
            # * 匹配任意多个字符
            正则模式 += '.*'
        elif 字符 == '?':
            # ? 匹配单个字符
            正则模式 += '.'
        elif 字符 == '#':
            # # 匹配单个数字
            正则模式 += '[0-9]'
        elif 字符 == '[':
            # 处理字符集 [abc] 或 [!abc]
            结束位置 = 模式.find(']', i)
            if 结束位置 == -1:
                正则模式 += re.escape(字符)
            else:
                字符集 = 模式[i:结束位置+1]
                # [!abc] 转为 [^abc]
                if len(字符集) > 2 and 字符集[1] == '!':
                    字符集 = '[^' + 字符集[2:]
                正则模式 += 字符集
                i = 结束位置
        else:
            # 普通字符需要转义
            正则模式 += re.escape(字符)
        
        i += 1
    
    # 完整匹配（从头到尾）
    正则模式 = '^' + 正则模式 + '$'
    
    return bool(re.match(正则模式, 文本, re.IGNORECASE))

def RGB(r, g=None, b=None) -> int:  
    """  
    将RGB颜色值或16进制颜色字符串转换为长整型数值（类似于VBA中的RGB函数）  
    
    Args:  
        r: 可以是以下格式之一：  
           - int: 红色分量 (0-255)，需要同时提供 g 和 b  
           - str: 16进制颜色字符串，如 "#F12121", "F12121", "#fff", "abc"  
        g (int, optional): 绿色分量 (0-255)  
        b (int, optional): 蓝色分量 (0-255)  
        
    Returns:  
        int: 长整型颜色值，格式为 B*65536 + G*256 + R  
        
    Examples:  
        RGB(255, 33, 33) -> 2171135  
        RGB("#F12121") -> 2171135  
    Raises:  
        ValueError: 当参数格式不正确时  
    """  
    # 如果 r 是字符串，解析16进制颜色  
    if isinstance(r, str):  
        # 移除开头的 # 号  
        hex_color = r.lstrip('#')  
        
        # 处理3位简写形式（如 #fff -> #ffffff）  
        if len(hex_color) == 3:  
            hex_color = ''.join([c*2 for c in hex_color])  
        
        # 验证长度  
        if len(hex_color) != 6:  
            raise ValueError(f"无效的16进制颜色格式: {r}，应为 #RRGGBB 或 #RGB 格式")  
        
        # 验证是否为合法的16进制字符  
        try:  
            # 解析16进制字符串  
            r_val = int(hex_color[0:2], 16)  
            g_val = int(hex_color[2:4], 16)  
            b_val = int(hex_color[4:6], 16)  
        except ValueError:  
            raise ValueError(f"无效的16进制颜色格式: {r}，包含非16进制字符")  
        
        return b_val * 65536 + g_val * 256 + r_val  
    
    # 如果是数值形式，需要提供所有三个参数  
    elif isinstance(r, int):  
        if g is None or b is None:  
            raise ValueError("使用RGB数值模式时，必须提供 r, g, b 三个参数")  
        
        # 验证范围  
        if not (0 <= r <= 255 and 0 <= g <= 255 and 0 <= b <= 255):  
            raise ValueError("RGB值必须在 0-255 范围内")  
        
        return b * 65536 + g * 256 + r  
    
    else:  
        raise ValueError(f"不支持的参数类型: {type(r)}，应为 int 或 str")  

#excel操作
_excelapp=None
def cleanup_com():
    try:
        pythoncom.CoUninitialize()
        #print("已清理COM资源")
    except Exception:
        pass

try:
    import pythoncom
    import atexit
    atexit.register(cleanup_com)
except Exception:
    pass

def kill_hidden_excel_et():
    """杀死无窗口的 Excel 和 ET 进程（简化版）"""
    try :
        import psutil
        import win32gui
        import win32process
    except ImportError:
        print("请安装 psutil 模块 uv add psutil")
        return False

    # 获取所有可见窗口的进程ID
    visible_pids = set()
    
    def callback(hwnd, _):
        if win32gui.IsWindowVisible(hwnd):
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
            visible_pids.add(pid)
        return True
    
    win32gui.EnumWindows(callback, None)
    
    # 查找并杀死无窗口的进程
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            name = proc.info['name'].upper()
            pid = proc.info['pid']
            
            if name in ["EXCEL.EXE", "ET.EXE"]:
                if pid not in visible_pids:
                    #print(f"杀死隐藏进程: {name} (PID: {pid})")
                    proc.kill()
                else:
                    pass
                    #print(f"保留进程: {name} (PID: {pid}) - 有可见窗口")
        except Exception:
            pass

    return True

def excel连接Com对象(强制重连=False, 清理隐藏进程=False):
    global _excelapp
    if _excelapp is None or 强制重连:
        if 清理隐藏进程:
            kill_hidden_excel_et()
        try:
            import win32com.client
            try:
                _excelapp = win32com.client.GetActiveObject("Ket.Application")  # WPS表格的应用程序名称
            except Exception:
                try:
                    _excelapp = win32com.client.GetActiveObject("Excel.Application")  # 尝试连接Excel
                except Exception:
                    raise Exception("未找到已打开的WPS或Excel应用程序")   
        except ImportError as e:
            if "win32com" in str(e):
                raise Exception("未安装pywin32库，请使用命令 'uv add pywin32' 安装")
        except Exception as e:
            print(f"连接Excel失败: {e}")
    return _excelapp
def excel创建新对象(可见=True):
    try:
        kill_hidden_excel_et()
        app=None
        import win32com.client
        try:
            app= win32com.client.Dispatch("Ket.Application")  # WPS表格的应用程序名称
        except Exception:
            try:
                app= win32com.client.Dispatch("Excel.Application")  # 尝试连接Excel
            except Exception:
                raise Exception("未找到WPS或Excel应用程序")   
        if app is None:
            raise Exception("未找到已打开的WPS或Excel应用程序")
        if 可见:
            app.Visible=True
        if app.Workbooks.Count==0:
            app.Workbooks.Add()
        return app
    except ImportError as e:
        if "win32com" in str(e):
            raise Exception("未安装pywin32库，请使用命令 'uv add pywin32' 安装")
        return None
# 不在导入模块时自动连接 Excel/WPS，避免 import hy127.vba 触发杀进程、弹窗或 COM 副作用。
excel = None

def list2D(数据):
    """
    将任意数据转换为规整的二维列表
    
    参数:
        数据: 可以是任意类型：None、单个值、元组、一维列表/元组、二维列表/元组 df 等
        
    返回:
        list[list]: 规整的二维列表，所有行具有相同的列数，不足部分用None补齐
    
    示例:
        list2D(5) -> [[5]]
        list2D([1, 2, 3]) -> [[1, 2, 3]]
        list2D([(1, 2), (3, 4)]) -> [[1, 2], [3, 4]]
        list2D([[1], [2, 3, 4]]) -> [[1, None, None], [2, 3, 4]]
    """
    # 处理 None 或空数据
    if 数据 is None:
        return [[]]
    
    # 转换元组为列表（递归处理嵌套元组）
    def 转为列表(obj):
        if isinstance(obj, tuple):
            return list(obj)
        return obj
    
    try: #处理df和series数据类型
        try:
            from .pandas_helper import df_to_list
        except ImportError:
            from pandas_helper import df_to_list
        import pandas as pd
        if isinstance(数据, (pd.DataFrame, pd.Series)):
            数据=df_to_list(数据)
    except Exception:
        pass

    数据 = 转为列表(数据)

    
    # 如果不是列表类型，说明是单个值
    if not isinstance(数据, (list, tuple)):
        return [[数据]]
    
    # 如果是空列表/元组
    if len(数据) == 0:
        return [[]]
    
    # 判断是否为二维结构
    # 检查第一个元素，如果是列表或元组，则认为是二维
    是二维 = isinstance(数据[0], (list, tuple))
    
    # 如果是一维列表/元组
    if not 是二维:
        return [list(数据) if isinstance(数据, tuple) else 数据]
    
    # 处理二维结构
    # 先将所有行转换为列表（处理元组的情况）
    临时结果 = []
    for 行 in 数据:
        if isinstance(行, (list, tuple)):
            临时结果.append(list(行) if isinstance(行, tuple) else 行)
        else:
            # 如果某一行是单个值，转为单元素列表
            临时结果.append([行])
    
    # 计算最大列数
    最大列数 = max(len(行) for 行 in 临时结果) if 临时结果 else 0
    
    # 如果最大列数为0，返回空行
    if 最大列数 == 0:
        return [[]]
    
    # 补齐所有行到相同列数
    结果 = []
    for 行 in 临时结果:
        补齐后的行 = 行 + [None] * (最大列数 - len(行))
        结果.append(补齐后的行)
    
    return 结果

def com日期时间矫正(数据, 指定列=None):
    """
    超高性能版本 - 适用于大数据量场景
    """

    import pywintypes
    from datetime import datetime
    数据=list2D(数据)

    if not 数据 or len(数据) == 0:
        return []
    
    第一行列数 = len(数据[0])
    指定列集合 = set(指定列) if 指定列 is not None else set(range(第一行列数))
    
    # 预先判断是否所有列都需要处理
    处理所有列 = 指定列集合 == set(range(第一行列数))
    
    if 处理所有列:
        # 所有列都处理，无需判断索引
        结果 = [
            [
                datetime.fromisoformat(元素.isoformat()) 
                if isinstance(元素, pywintypes.TimeType) 
                else 元素
                for 元素 in 行
            ]
            for 行 in 数据
        ]
    else:
        # 部分列处理
        结果 = [
            [
                datetime.fromisoformat(元素.isoformat()) 
                if i in 指定列集合 and isinstance(元素, pywintypes.TimeType) 
                else 元素
                for i, 元素 in enumerate(行)
            ]
            for 行 in 数据
        ]
    
    return 结果

def excel读取到df(地址=None,表名=None,带表头=True,文件路径=None,自动创建Excel对象=False,扩展到最大行=True ):
    """
    通过Excel COM对象读取Excel数据
    参数:
        文件路径: Excel文件路径，为空则使用当前活动工作簿
        表名: 工作表名称，为空则使用当前活动工作表
        地址: 单元格地址范围，为空则使用已使用区域，否则从该地址开始读取到最大有效数据行
    
    返回:
        DataFrame: pandas数据帧对象
    """
    global excel
    if 自动创建Excel对象 and excel is None:
            excel=excel创建新对象(False)
    try:
        import pandas as pd
        arr=excel读取到list(地址,表名,文件路径)
        if 带表头:
             return pd.DataFrame(arr[1:],columns=arr[0])
        else:
            return pd.DataFrame(arr)
    except ImportError as e:
        if "pandas" in str(e):
            raise Exception("未安装pandas库，请使用命令 'uv add pandas' 安装")
        else:
            raise e
    except Exception as e:
        raise Exception(f"读取数据时发生错误: {e}")
#end def    
def excel读取到list(单元格地址=None,表名=None,文件路径=None,自动创建Excel对象=False,扩展到最大行=True )->list[list]:
    """
    通过Excel COM对象读取Excel数据
    参数:
        文件路径: Excel文件路径，为空则使用当前活动工作簿
        表名: 工作表名称，为空则使用当前活动工作表
        地址: 单元格地址范围，为空则使用已使用区域，否则从该地址开始读取到最大有效数据行
    
    返回:
        list: 二维列表形式的Excel数据
    """
    try:
        # 获取Excel应用程序对象
        global excel
        excel_app = excel
        if 自动创建Excel对象 and excel_app is None:
            excel_app=excel创建新对象(False)
        if '!' in 单元格地址:
            表名,单元格地址=单元格地址.split('!')
            #表名删除单引号
            表名=表名.replace("'","")
        # 获取工作簿
        if 文件路径:
            workbook = excel获取或打开工作簿(文件路径,excel_app)
        else:
            workbook = excel_app.ActiveWorkbook
        excel= excel_app
        # 获取工作表
        if 表名:
            worksheet = workbook.Worksheets(表名)
        else:
            worksheet = workbook.ActiveSheet
            
        # 获取数据范围
        if 单元格地址:
             # 确定起始单元格
            if type(单元格地址) is str:
              
                start_range = worksheet.Range(单元格地址)
                if 扩展到最大行:
                    # 获取最后一行
                    if start_range.Rows.Count > 1: #本身多行的不往下扩展
                        last_row =start_range.Rows.Count+start_range.Row
                    else:
                        last_row = worksheet.UsedRange.Rows.Count + worksheet.UsedRange.Row - 1
                else:
                    last_row =start_range.Rows.Count+start_range.Row
                
                # 构造完整范围
                data_range = start_range.GetResize(max(last_row-start_range.Row+1,1))
            else:
                data_range=单元格地址
                worksheet=data_range.Worksheet
           
        else:
            # 如果没有提供地址，则使用已使用区域
            data_range = worksheet.UsedRange
            
        # 读取数据并转换为二维列表
        if data_range.Rows.Count > 0 and data_range.Columns.Count > 0:
            # 获取数据值
            values = data_range.Value
            # 处理只有一行的情况
            if data_range.Rows.Count == 1:
                if data_range.Columns.Count == 1:
                    # 只有一个单元格
                    data = [[values]] if values is not None else [[]]
                else:
                    # 只有一行多列
                    data = [list(values) if values is not None else [] * data_range.Columns.Count]
            else:
                # 多行情况 - 使用列表推导式简化代码
                data = [list(row) if isinstance(row, tuple) else [row] for row in values]
            
            # 从后往前检查并删除连续的空行
            while data and all(cell == "" or cell is None for cell in data[-1]):
                data.pop()
                
            # 如果所有行都被删除了，返回包含一个空行的列表
            if not data:
                data = [[]]
                
            return com日期时间矫正(data)
        else:
            return [[]]
            
    except Exception as e:
        # 发生异常时返回空数组
        return [[]]
#end def
def excel获取或打开工作簿(文件路径, excel_app=None):
    """
    检查Excel文件是否已经打开，如果已打开则返回对应的工作簿对象，否则打开文件
    """
    import os
    
    if excel_app is None:
        global excel
        excel_app = excel
    
    # 检查文件是否已经打开
    for workbook in excel_app.Workbooks:
        if workbook.FullName.lower() == 文件路径.lower():
            return workbook
    
    # 如果没有找到已打开的工作簿，则打开文件
    return excel_app.Workbooks.Open(文件路径)
def excel写入(数据, 单元格地址="A1",表名=None, 文件路径=None,清空区域下方=True,加边框=True,自适应列宽=True):
    """
    通过Excel COM对象将数据写回到Excel
    
    参数:
        数据: 要写入的二维列表数据
        单元格地址: 起始单元格地址，为空则从A1开始 也可以转Range对象
        表名: 工作表名称，为空则使用当前活动工作表
        文件路径: Excel文件路径，为空则使用当前活动工作簿
        清空区域下方: 是否清空区域下方的行，默认为True
        加边框 : 添加边框线 默认为trye
    返回值: 
        返回写入的Range对象
    """

    try:
        global excel
        # 获取Excel应用程序对象
        if not type(单元格地址) is str:
            excel_app=单元格地址.Application #传入的是对象则直接使用
        else:            
            if  excel is None:
                excel=excel创建新对象(False)
            excel_app = excel
              
        if '!' in 单元格地址:
            表名,单元格地址=单元格地址.split('!')
            #表名删除单引号
            表名=表名.replace("'","")   

        # 获取工作簿
        if 文件路径:
            import os
            if os.path.exists(文件路径):
                workbook = excel获取或打开工作簿(文件路径)
            else:
                # 文件不存在，创建新的工作簿
                workbook = excel_app.Workbooks.Add()
                # 保存为指定路径
                workbook.SaveAs(文件路径)
        else:
            workbook = excel_app.ActiveWorkbook
            
        excel=excel_app    
        # 获取工作表
        if 表名:
            try:
                worksheet = workbook.Worksheets(表名)
            except Exception:
                # 工作表不存在，创建新的工作表
                worksheet = workbook.Worksheets.Add()
                worksheet.Name = 表名
        else:
            worksheet = workbook.ActiveSheet
        # 确定起始单元格
        if 单元格地址:
            if type(单元格地址) is str:
                start_cell = worksheet.Range(单元格地址)
            else:
                start_cell=单元格地址
                worksheet=start_cell.Worksheet
        else:
            start_cell = worksheet.Range("A1")
       
            # 获取当前区域下方的行数    
        # 计算数据范围
        数据=list2D(数据)
        if 数据:
            rows = len(数据)
            cols = len(数据[0]) if 数据[0] else 1
            # 获取目标范围
            target_range = start_cell.GetResize(RowSize=rows, ColumnSize=cols)
            if 清空区域下方:
            # 获取当前区域下方的行数
                last_row = worksheet.UsedRange.Rows.Count + worksheet.UsedRange.Row - 1
                crng=start_cell.GetResize(max(last_row-target_range.Row+1,1),cols)
                try:
                    crng.ClearContents()
                except Exception:
                    try:
                        crng._ClearContents()
                    except Exception:
                        pass
            # 将数据写入Excel
            target_range.Value = 数据
            if 加边框:
                bd=target_range.Borders
                bd.LineStyle = 1
                bd.Weight = 2
            if 自适应列宽:
                target_range.EntireColumn.AutoFit() #列宽自适应
        # 如果提供了文件路径，保存并关闭工作簿
        if 文件路径:
            try:
                workbook.Save()
                if not excel_app.Visible: #不可见状态下就直接关闭
                    workbook.Close()
                    excel_app.Quit()
            except Exception:
               pass
           
            
        return  target_range    
    except Exception as e:
        raise Exception(f"写入Excel时发生错误: {e}")
#end def
def excel读取到df_pandas(地址=None, 表名=None, 带表头=True, 文件路径=None):
    """
    使用pandas.read_excel读取Excel数据到DataFrame
    
    参数:
        地址: 单元格地址范围，支持以下几种形式：
              - A:F 无行号模式，读取所有行的A到F列
              - A1:F1 行号相同，自动向下扩展读取数据
              - A1:F10 指定具体区域
        表名: 工作表名称，为空则使用第一个工作表
        带表头: 是否将第一行作为列名，默认True
        文件路径: Excel文件路径，为空则抛出异常
        
    返回:
        DataFrame: pandas数据帧对象
    """
    try:
        import pandas as pd
    except ImportError as e:
        if "pandas" in str(e):
            raise Exception("未安装pandas库，请使用命令 'uv add pandas' 安装")
        else:
            raise e
    
    if not 文件路径:
        raise Exception("使用pandas方式读取必须提供文件路径")
    
    # 处理"表名!地址"的形式
    if 地址 and '!' in 地址:
        表名, 地址 = 地址.split('!')
        表名 = 表名.replace("'", "")
    
    # 解析地址
    usecols = None
    skiprows = None
    nrows = None
    
    if 地址:
        if ':' in 地址:
            start_col_row, end_col_row = 地址.split(':')
            
            # 提取列和行信息
            start_col = ''.join(filter(str.isalpha, start_col_row))
            start_row = ''.join(filter(str.isdigit, start_col_row))
            
            end_col = ''.join(filter(str.isalpha, end_col_row))
            end_row = ''.join(filter(str.isdigit, end_col_row))
            
            # 设置usecols参数
            usecols = f"{start_col}:{end_col}"
            
            # 处理行号情况
            if start_row and end_row:
                if start_row == end_row:
                    # 行号相同，从该行开始向下读取所有数据
                    skiprows = int(start_row) - 1
                else:
                    # 指定了具体的行范围
                    skiprows = int(start_row) - 1
                    nrows = int(end_row) - int(start_row) + 1
        else:
            # 单个单元格或列
            col = ''.join(filter(str.isalpha, 地址))
            row = ''.join(filter(str.isdigit, 地址))
            
            usecols = col
            if row:
                skiprows = int(row) - 1
    
    # 读取Excel
    try:
        kwargs = {
            'header': 0 if 带表头 else None,
            'sheet_name': 表名 if 表名 else 0
        }
        
        if usecols:
            kwargs['usecols'] = usecols
        if skiprows is not None:
            kwargs['skiprows'] = skiprows
        if nrows is not None:
            kwargs['nrows'] = nrows
            
        df = pd.read_excel(文件路径, **kwargs)
        return df
    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {e}")
#end def excel读取到df_pandas
def excel读取到list_pandas(单元格地址=None,表名=None,文件路径=None)->list[list]:
    """
    通过pandas读取Excel数据到list
    参数:
        文件路径: Excel文件路径，为空则使用当前活动工作簿
        表名: 工作表名称，为空则使用当前活动工作表
    """ 
   
    df=excel读取到df_pandas(单元格地址,表名=表名,文件路径=文件路径,带表头=False)
    return df.values.tolist() #没有表头直接转为list

    pass
#end def
def excel写入_pandas(数据, 单元格地址="A1",表名=None, 文件路径=None, 是否绘制表格边框=True, 清空下方区域=True):
    """
    将二维列表写入Excel指定区域，使用行列坐标直接定位
    
    参数:
        二维列表: 要写入的数据，格式为 [[行1], [行2], ...]
        文件路径: Excel文件路径
        工作表名: 目标工作表名称
        是否绘制表格边框: 是否为表格添加边框，默认为 False
        清空下方区域: 是否清空指定区域下方的数据，默认为 True
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side
    数据 = list2D(数据)
    if '!' in 单元格地址:
        表名,单元格地址=单元格地址.split('!')
        #表名删除单引号
        表名=表名.replace("'","")   
    # 加载工作簿
        # 检查文件是否存在，不存在则新建
    from openpyxl import load_workbook, Workbook
    
    if os.path.exists(文件路径):
        wb = load_workbook(文件路径)
    else:
        wb = Workbook()
        # 保存新创建的工作簿以便后续加载
        if 表名:
           #修改活动工作表表名为表名
           wb.active.title = 表名
        wb.save(文件路径)
        # 重新加载以确保正确初始化
        wb = load_workbook(文件路径)
    
    # 确保工作表存在
    if 表名 and 表名 in wb.sheetnames:
        ws = wb[表名]
    elif 表名:
        ws = wb.create_sheet(表名)
    else:
        ws = wb.active

    ws = wb[表名]
    
    # 定义边框样式
    边框样式 = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    起始行, 起始列 = excel地址转行列索引(单元格地址)
    
    # 如果需要清空下方区域，则先清空
    if 清空下方区域:
        # 清空数据区域
        for row_idx in range(起始行, ws.max_row + 1):
            for col_idx in range(起始列, 起始列 + len(数据[0]) if 数据 else 0):
                ws.cell(row=row_idx, column=col_idx).value = None
    
    # 写入数据，直接使用行列索引
    for i, 行 in enumerate(数据):
        for j, 值 in enumerate(行):
            cell = ws.cell(row=起始行 + i, column=起始列 + j,value=值)
            # 如果需要绘制边框，则应用边框样式
            if 是否绘制表格边框:
                cell.border = 边框样式
    # 保存文件
    wb.save(文件路径)
    return True
#end def
def excel地址转行列索引(地址):
    地址 = 地址.strip().upper()
    匹配结果 = re.match(r'^([A-Z]+)(\d+)$', 地址)
    列字母 = 匹配结果.group(1)
    行号 = int(匹配结果.group(2))
    列索引 = 0
    for 字符 in 列字母:
        列索引 = 列索引 * 26 + (ord(字符) - ord('A') + 1)
    return 行号, 列索引
#end def


try:#拼音组件
    from pypinyin import lazy_pinyin,Style as PYStyle
except ImportError:
    lazy_pinyin = None
    PYStyle = None

try:
    import pandas as pd
except Exception:
    pass

def df行拼音排序回调(数据系列):
   
    """
    将中文转换为拼音用于排序的回调函数
    
    参数:
        数据系列: pandas Series对象
    返回:
        转换为拼音后的Series对象
    示范:
        df.sort_values('姓名', key=df拼音排序回调)
    """
    
    if lazy_pinyin is None or pd is None:
        raise ImportError("使用拼音排序需要安装 pypinyin 和 pandas")
    return 数据系列.apply(lambda x: ''.join(lazy_pinyin(str(x))) if pd.notnull(x) else x)

def df列拼音排序回调(列索引系列):
    """列索引拼音排序回调"""
    if lazy_pinyin is None or pd is None:
        raise ImportError("使用拼音排序需要安装 pypinyin 和 pandas")
    return 列索引系列.to_series().apply(
        lambda x: ''.join(lazy_pinyin(str(x))) if pd.notnull(x) else x
    )

def z拼音首字母(x):
    if lazy_pinyin is None:
        raise ImportError("使用拼音功能需要安装 pypinyin")
    return ''.join(lazy_pinyin(str(x),style=PYStyle.FIRST_LETTER)) if not x is None else x

def z拼音全拼(x):
    if lazy_pinyin is None:
        raise ImportError("使用拼音功能需要安装 pypinyin")
    return ''.join(lazy_pinyin(str(x))) if not x is None else x

# 新增二维列表拼音排序回调函数
def list拼音排序回调(列索引=0):
    """
    为二维列表创建按拼音排序的回调函数，可指定列索引
    参数:
        列索引: 要排序的列索引，默认为0（第一列）
    返回:
        排序回调函数
    示范:
        sorted(数据, key=list拼音排序回调(1))  # 按第二列拼音排序
    """
    def 排序函数(行):
        # 获取指定列的值
        值 = 行[列索引] if 列索引 < len(行) else None
        # 转换为拼音字符串用于排序
        return ''.join(lazy_pinyin(str(值))) if 值 is not None else ''
    return 排序函数


def safe_used_range(工作表=None):
    """
    通过最大行来确定使用区域，系统使用区域有时候不准
    
    参数:
        工作表: Excel工作表对象，为空则使用当前活动工作表
        
    返回:
        Range: Excel区域对象，包含从A1到最后一个非空单元格的区域
    """
    try:
        global excel
        if 工作表 is None:
            if excel is None:
                excel = excel连接Com对象()
            工作表 = excel.ActiveSheet
        
        # 计算最后一个非空行号
        最后行 = 工作表.Cells.Find("*", 工作表.Cells(1, 1), -4163, -4162, 1, 2).Row
        # 计算最后一个非空列号
        最后列 = 工作表.Cells.Find("*", 工作表.Cells(1, 1), -4163, -4162, 2, 2).Column
        
        # 返回从A1到最后一个非空单元格的区域
        return 工作表.Range(工作表.Cells(1, 1), 工作表.Cells(最后行, 最后列))
        
    except Exception as e:
        # 如果查找失败，返回已使用区域
        return 工作表.UsedRange

def max_range(区域, 参考列=""):
    """
    某一起始行区域往下延展的最大区域，默认按所有列最大行，也可指定参考列
    
    参数:
        区域: Excel区域对象或字符串地址（如"A1:D1"）
        参考列: 可选，指定参考列（如"D"），为空则按所有列最大行
        
    返回:
        Range: 向下扩展后的区域对象
        
    示例:
        max_range([A1:D1], "D").Select
        max_range("A1:D1", "D").Select
    """
    try:
        global excel
        if excel is None:
            excel = excel连接Com对象()
        
        # 处理字符串类型的区域参数
        if isinstance(区域, str):
            # 如果是字符串，使用当前活动工作表的Range方法获取区域对象
            工作表 = excel.ActiveSheet
            区域 = 工作表.Range(区域)
            
        工作表 = 区域.Worksheet
        
        if 参考列 == "":
            # 整体往下扩展到最后行
            扩展区域 = 区域.Resize(工作表.Rows.Count - 区域.Row + 1, 区域.Columns.Count)
        else:
            # 参考列往下扩展到最后行
            参考列索引 = excel地址转行列索引(参考列 + "1")[1]
            扩展区域 = 区域.EntireRow.Cells(1, 参考列索引).Resize(工作表.Rows.Count - 区域.Row + 1)
        
        # 计算当前行下方区域的最后一行
        # -4123: xlFormulas（包含隐藏行），-4163: xlValues（忽略隐藏行）
        最后行 = 扩展区域.Find("*", 扩展区域.Cells(1, 1), -4123, -4162, 1, 2).Row
        
        # 返回调整大小后的区域
        return 区域.Resize(最后行 - 区域.Row + 1)
        
    except Exception as e:
        # 如果查找失败，返回原始区域
        return 区域

def lastRow(区域):
    """
    返回区域所在列的最后行数
    
    参数:
        区域: Excel区域对象或字符串地址（如"A1"）
        
    返回:
        int: 区域所在列的最后一行行号
    """
    try:
        global excel
        if excel is None:
            excel = excel连接Com对象()
        
        # 处理字符串类型的区域参数
        if isinstance(区域, str):
            # 如果是字符串，使用当前活动工作表的Range方法获取区域对象
            工作表 = excel.ActiveSheet
            区域 = 工作表.Range(区域)
            
        工作表 = 区域.Worksheet
        最大行 = 工作表.Rows.Count
        
        # 使用End(xlUp)方法找到最后一个非空单元格
        最后行单元格 = 工作表.Cells(最大行, 区域.Column).End(-4162)  # -4162: xlUp
        
        return 最后行单元格.Row
        
    except Exception as e:
        # 如果出错，返回最大行
        return 最大行


# 函数别名定义
z四舍五入 = round
z转数值 = val
z字符串转日期时间 = str_to_datetime
z字符串转日期 = str_to_date
z转日期数值 = cdate
z转日期时间 = to_datetime
z转日期 = to_date
z判断取值 = IIF
z是空值 = isEmpty
z模糊匹配 = like
z三原色=RGB

z数据框行拼音排序回调=df行拼音排序回调
z数据框列拼音排序回调=df列拼音排序回调
z列表拼音排序回调=list拼音排序回调

# 新增Excel区域操作函数别名
z安全使用区域=safe_used_range
z最大扩展区域=max_range
z最大行号=lastRow


def _test():
    try:
        from .log_helper import logtable
    except ImportError:
        from log_helper import logtable
    logtable(list2D([[1,2,3,4,5],[],['a']]))
    excel窗口置顶(True)
    d=excel读取到list(单元格地址="A1:G1")
    logtable(d)
    sht=excel.Sheets.Add()
    sht.Range("A1").Value="test"
    sht.Range("A2").Select()
    sht.Range("A1").Interior.Color = RGB("#07C160")
    pass

if __name__ == "__main__":
    pass

